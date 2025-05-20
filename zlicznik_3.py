from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl

# Ścieżka do pliku Excel
excel_file = 'EXCELTram Data.xlsx'

# Wczytanie pliku Excel
wb = openpyxl.load_workbook(excel_file)
sheet = wb.active

# Pętla po wierszach w arkuszu, zaczynając od drugiego wiersza (min_row=2)
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    url = row[4].value  # Komórka E to 5. kolumna (indeks 4)
    if url:  # Sprawdzenie, czy URL nie jest pusty
        print(f'Analizowanie URL: {url}')
        try:
            # Wysyłanie żądania HTTP do strony i pobranie jej zawartości
            response = urlopen(url)
            html = response.read()
            soup = BeautifulSoup(html, 'html.parser')
            # Znalezienie wszystkich elementów li z określoną klasą
            li_elements = soup.find_all('li', class_='timetable-departures-entry')
            # Zliczenie elementów
            count = 0
            for li in li_elements:
                aria_label = li.find('a')['aria-label'] if li.find('a') else ''
                if aria_label.startswith('Tramwaj'):
                    count += 1
            print(f'Liczba Tramwaji: {count}')
            # Zapisanie wyniku do kolumny H w tym samym wierszu
            sheet.cell(row=row[0].row, column=8, value=count)  
        except Exception as e:
            print(f'Błąd podczas przetwarzania URL: {url}, {e}')
            sheet.cell(row=row[0].row, column=8, value='Błąd') 
    else:
        print(0)
        sheet.cell(row=row[0].row, column=8, value=0)  

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    url = row[5].value  # Komórka F to 6. kolumna (indeks 5)
    if url:  
        print(f'Analizowanie URL: {url}')
        try: 
            response = urlopen(url)
            html = response.read()
            soup = BeautifulSoup(html, 'html.parser')
            li_elements = soup.find_all('li', class_='timetable-departures-entry')
            count = 0
            for li in li_elements:
                aria_label = li.find('a')['aria-label'] if li.find('a') else ''
                if aria_label.startswith('Tramwaj'):
                    count += 1
            print(f'Liczba Tramwaji: {count}')
            # Zapisanie wyniku do kolumny I w tym samym wierszu
            sheet.cell(row=row[0].row, column=9, value=count)  
        except Exception as e:
            print(f'Błąd podczas przetwarzania URL: {url}, {e}')
            sheet.cell(row=row[0].row, column=9, value='Błąd') 
    else:
        print(0)
        sheet.cell(row=row[0].row, column=9, value=0) 

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    url = row[6].value 
    if url: 
        print(f'Analizowanie URL: {url}')
        try:
            response = urlopen(url)
            html = response.read()
            soup = BeautifulSoup(html, 'html.parser')
            li_elements = soup.find_all('li', class_='timetable-departures-entry')
            count = 0
            for li in li_elements:
                aria_label = li.find('a')['aria-label'] if li.find('a') else ''
                if aria_label.startswith('Tramwaj'):
                    count += 1
            print(f'Liczba Tramwaji: {count}')
            sheet.cell(row=row[0].row, column=10, value=count) 
        except Exception as e:
            print(f'Błąd podczas przetwarzania URL: {url}, {e}')
            sheet.cell(row=row[0].row, column=10, value='Błąd') 
    else:
        print(0)
        sheet.cell(row=row[0].row, column=10, value=0) 

# Zapisanie zmian w pliku Excel
wb.save(excel_file)
wb.close()
